#!/usr/bin/env python3
"""
Script to convert Wayground quiz Markdown to Excel format.
"""

import re
from openpyxl import Workbook
from pathlib import Path

def parse_quiz_markdown(md_file):
    """Parse the markdown quiz file and extract questions."""
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    questions = []
    
    # Split by question headers (##)
    question_blocks = re.split(r'^##\s+(.+)$', content, flags=re.MULTILINE)
    
    # Skip the first element (content before first question)
    for i in range(1, len(question_blocks), 2):
        if i + 1 >= len(question_blocks):
            break
            
        question_text = question_blocks[i].strip()
        question_content = question_blocks[i + 1].strip()
        
        # Skip if it's not a question (e.g., section headers)
        if question_text.startswith('#') or 'Question Distribution' in question_text or 'Topic Coverage' in question_text or 'Notes' in question_text:
            continue
        
        # Extract options
        options = []
        correct_answer_nums = []
        
        # Find all options (lines starting with - ✅ or - ❌)
        option_pattern = r'^-\s+([✅❌])\s+(.+)$'
        option_num = 0
        for line in question_content.split('\n'):
            match = re.match(option_pattern, line.strip())
            if match:
                option_num += 1
                is_correct = match.group(1) == '✅'
                option_text = match.group(2).strip()
                options.append(option_text)
                if is_correct:
                    correct_answer_nums.append(option_num)
        
        # Validate max 4 options
        if len(options) > 4:
            print(f"Warning: Question '{question_text[:50]}...' has {len(options)} options (max 4). Truncating to first 4.")
            options = options[:4]
            correct_answer_nums = [n for n in correct_answer_nums if n <= 4]
        
        # Extract explanation
        explanation_match = re.search(r'\*\*Explanation\*\*:\s*(.+?)(?=\n\n|\n##|$)', question_content, re.DOTALL)
        explanation = explanation_match.group(1).strip() if explanation_match else ''
        
        # Determine question type
        if len(options) == 2 and ('True' in options[0] or 'False' in options[0]):
            question_type = 'True/False'
            # True/False should have single correct answer
            correct_answer = str(correct_answer_nums[0]) if correct_answer_nums else None
        else:
            question_type = 'Multiple Choice'
            # Format correct answer: single number only
            # Note: Only using single correct answers per question for Wayground compatibility
            if len(correct_answer_nums) == 1:
                correct_answer = str(correct_answer_nums[0])
            elif len(correct_answer_nums) > 1:
                # If multiple correct answers found, use the first one and warn
                print(f"Warning: Question '{question_text[:50]}...' has {len(correct_answer_nums)} correct answers. Using only the first one.")
                correct_answer = str(correct_answer_nums[0])
            else:
                correct_answer = None
        
        questions.append({
            'question': question_text,
            'type': question_type,
            'options': options,
            'correct_answer': correct_answer,
            'explanation': explanation
        })
    
    return questions

def create_excel_file(questions, output_file):
    """Create Excel file from questions."""
    wb = Workbook()
    ws = wb.active
    # Keep default sheet name "Sheet" to match Wayground format
    
    # Header row - matching Wayground template format
    headers = ['Question Text', 'Question Type', 'Option 1', 'Option 2', 'Option 3', 'Option 4', 'Option 5', 'Correct Answer', 'Time in seconds', 'Image Link', 'Answer explanation']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
    
    # Add questions
    for row_num, q in enumerate(questions, 2):
        ws.cell(row=row_num, column=1, value=q['question'])  # Question Text
        ws.cell(row=row_num, column=2, value=q['type'])  # Question Type
        
        # Add options (columns 3-7)
        for opt_num, option in enumerate(q['options'], 1):
            ws.cell(row=row_num, column=2 + opt_num, value=option)
        
        # Fill remaining option columns with None if needed (max 4 options, but Excel has 5 columns)
        for opt_num in range(len(q['options']) + 1, 6):
            ws.cell(row=row_num, column=2 + opt_num, value=None)
        
        ws.cell(row=row_num, column=8, value=q['correct_answer'])  # Correct Answer (number or comma-separated)
        ws.cell(row=row_num, column=9, value=25)  # Time in seconds (default 25)
        ws.cell(row=row_num, column=10, value=None)  # Image Link (empty)
        ws.cell(row=row_num, column=11, value=q['explanation'])  # Answer explanation
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"Excel file created: {output_file}")

def main():
    md_file = Path('quarto/sessions/12_history-of-english/quiz/wayground-history-of-english-quiz.md')
    output_file = Path('quarto/sessions/12_history-of-english/quiz/wayground-history-of-english-quiz.xlsx')
    
    if not md_file.exists():
        print(f"Error: Markdown file not found: {md_file}")
        return
    
    questions = parse_quiz_markdown(md_file)
    print(f"Parsed {len(questions)} questions")
    
    create_excel_file(questions, output_file)

if __name__ == '__main__':
    main()
